
import tkinter.messagebox
import tkinter as tk
import random
import string
import openpyxl
import pymysql.cursors
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import DataBase
import pandas as pd

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        input_window = tk.Tk()
        input_window.title("Ввод списка")

        self.db_name = None
        self.table_name = None
        self.title("Создание базы данных")

        self.db_label = tk.Label(self, text="Имя базы данных:")
        self.db_label.pack()
        self.db_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.db_entry.pack()

        self.table_label = tk.Label(self, text="Имя таблицы:")
        self.table_label.pack()
        self.table_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.table_entry.pack()

        self.list_label = tk.Label(input_window, text="Введите 85 отрицательных и положительных элементов:")
        self.list_label.pack()
        self.list_entry = tk.Entry(input_window, width=50)
        self.list_entry.pack()

        self.title("Записи из таблицы MySQL")
        self.records_text = tk.Text(self)
        self.records_text.pack()

        self.file1_label1 = tk.Label(self, text="Имя файла эксель:")
        self.file1_label1.pack()
        self.file1_entry1 = tk.Entry(self, textvariable=tk.StringVar())
        self.file1_entry1.pack()

        self.db_label.pack(padx=50, pady=2)
        self.table_label.pack(padx=50, pady=2)
        self.file1_label1.pack(padx=50, pady=2)

    def create_database(self):
        self.db_name = self.db_entry.get()
        self.table_name = self.table_entry.get()

        db = DataBase.DataBase(self.db_name, self.table_name)
        db.check_db()
        db.check_table()
        self.mysql_tb()

    def rand_alf_sample(self):
        gl = 'aeyuio'
        list_alf = [str(i) for i in self.sample_120_entry.get().split()]
        #list_alf = [random.choice(string.ascii_lowercase) for _ in range(120)]
        count_gl = 0
        count_sogl = 0
        for i in list_alf:
            if i in gl:
                count_gl += 1
            else:
                count_sogl += 1
        return list_alf, count_gl, count_sogl

    def separate_numbers(self, numbers):
        positive_numbers = [num for num in numbers if num > 0]
        negative_numbers = [num for num in numbers if num < 0]
        return positive_numbers, negative_numbers


    def inp_list(self):
        numbers = self.list_entry.get().split()
        positive_numbers, negative_numbers = self.separate_numbers(numbers)
    def mysql_tb(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        with connection.cursor() as cursor:
            sql = f"SELECT * FROM {self.table_name}"
            cursor.execute(sql)
            records = cursor.fetchall()

            for record in records:
                self.records_text.insert(tk.END, f"{record}\n")

    def save_result(self):
        input_list, even_indices_str = self.inp_list()
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"INSERT INTO {db1.name_tb} (list_ind, indexi) VALUES (%s, %s)",
                               (str(input_list), str(even_indices_str)))
                connection.commit()

                cursor.execute(f"SELECT * FROM {db1.name_tb}")
                print(cursor.fetchall()[-1])

        except pymysql.err.DataError as e:
            print('Ошибка с данными:', e)

        except pymysql.err.DatabaseError as e:
            print(e)

    def list_tb(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        cursor = connection.cursor()
        tb_in_db = "SHOW TABLES;"
        cursor.execute(tb_in_db)
        tables = cursor.fetchall()

        table_list = [table[0] for table in tables]
        table_list_str = "\n".join(table_list)

        tkinter.messagebox.showinfo("Список таблиц", table_list_str)


    def save_to_excel(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        try:
            new_df = pd.read_sql("SELECT * FROM " + self.table_name, connection)
            wb = openpyxl.Workbook()
            ws = wb.active

            for r in dataframe_to_rows(new_df, index=False, header=True):
                ws.append(r)

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            file1 = self.file1_entry1.get()
            wb.save(file1)
            print(new_df)

            tkinter.messagebox.showinfo("Импорт в эксель", file1)

        except pymysql.err.DatabaseError as e:
            print(e)
        return


app = App()

create_button1 = tk.Button(app, text="Создать БД", command=app.create_database)
create_button1.pack()

button = tk.Button(app, text="Список", command=app.inp_list)
button.pack()

create_button = tk.Button(app, text="Создать запись", command=app.save_result)
create_button.pack()

list_button = tk.Button(app, text="Показать список таблиц", command=app.list_tb)
list_button.pack()

excel_button = tk.Button(app, text="Импорт в эксель", command=app.save_to_excel)
excel_button.pack()


app.mainloop()